"""Shared utilities for report generation."""

import csv
from datetime import date
from io import BytesIO, StringIO

import openpyxl
from flask_sqlalchemy.query import Query
from sqlalchemy.orm import joinedload

from backend.models import PayrollSettings, Resident, Role, TimeEntry, db
from backend.type_defs import (
    ResidentData,
    ResidentEntryDict,
    ResidentID,
    ResidentSummaryDict,
    TimeEntries,
)


def build_entries_query(
    start_date: date, end_date: date, resident_id: ResidentID
) -> Query:
    """Build query for entries within date range, filtered by resident.

    Uses an inner join on Role, so entries whose role was deleted (role_id set
    to NULL by the ON DELETE SET NULL constraint) are intentionally excluded —
    they carry no role context and cannot contribute to overtime calculations.
    Call team roles (is_call_team=True) are also excluded.
    Resident and role relationships are eagerly loaded to avoid N+1 queries.
    """
    query = (
        TimeEntry.query.join(Role)
        .options(joinedload(TimeEntry.resident), joinedload(TimeEntry.role))
        .filter(
            TimeEntry.date >= start_date,
            TimeEntry.date <= end_date,
            Role.is_call_team.isnot(True),
        )
    )
    if resident_id:
        query = query.filter(TimeEntry.resident_id == resident_id)
    return query


def get_resident_name(resident_id: ResidentID) -> str | None:
    """Look up resident name by ID."""
    if not resident_id:
        return None
    resident = db.session.get(Resident, resident_id)
    return resident.name if resident else None


def aggregate_entries_by_resident(entries: TimeEntries) -> ResidentData:
    """
    Aggregate time entries by resident.

    Returns:
        Dictionary mapping resident_id (int) to their entries and total overtime.
        Example: {42: {"name": "John Doe", "entries": [...], "total_overtime": 2.5}}
    """
    resident_data: ResidentData = {}
    for entry in entries:
        resident: Resident = entry.resident
        role: Role = entry.role  # type: ignore[assignment]  # inner join guarantees non-null
        res_id: int = resident.id
        if res_id not in resident_data:
            resident_data[res_id] = ResidentSummaryDict(
                name=resident.name,
                entries=[],
                total_overtime=0.0,
            )

        overtime = entry.overtime_hours
        resident_data[res_id]["entries"].append(
            ResidentEntryDict(
                date=entry.date.strftime("%Y-%m-%d"),
                role=role.name,
                exit_time=entry.exit_time.strftime("%H:%M") if entry.exit_time else "",
                overtime=overtime,
            )
        )
        resident_data[res_id]["total_overtime"] += overtime

    return resident_data


def generate_csv_content(entries: TimeEntries) -> str:
    """Generate detailed CSV content from time entries."""
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Resident", "Role", "Exit Time", "Overtime Hours"])

    for entry in entries:
        resident: Resident = entry.resident
        role: Role = entry.role  # type: ignore[assignment]  # inner join guarantees non-null
        writer.writerow(
            [
                entry.date.strftime("%Y-%m-%d"),
                resident.name,
                role.name,
                entry.exit_time.strftime("%H:%M") if entry.exit_time else "",
                f"{entry.overtime_hours:.2f}",
            ]
        )

    return output.getvalue()


def generate_payroll_xlsx(
    resident_data: ResidentData,
    start_date: date,
    end_date: date,
    settings: PayrollSettings,
) -> bytes:
    """
    Generate a Lawson/UPHS payroll export spreadsheet (A..AB).

    Residents are included even when lawson_id is missing.
    Dates are written as true Excel dates; code-like fields are forced to text.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    layout = settings.layout
    c = layout.columns

    ws.append(layout.headers)

    codes = settings.export_codes()
    note = settings.note_for(start_date)
    text_cols = settings.export_text_cols()

    # Precompute constants once
    base_row: list[object | None] = [None] * layout.n_cols
    base_row[c["program"]] = codes["program"]
    base_row[c["company"]] = codes["company"]
    base_row[c["batch"]] = codes["batch"]
    base_row[c["pay_code"]] = codes["pay_code"]
    base_row[c["transdate"]] = end_date  # date object
    base_row[c["dept"]] = codes["dept"]
    base_row[c["expense"]] = codes["expense"]
    base_row[c["acct_unit"]] = codes["acct_unit"]
    base_row[c["note"]] = note

    hire_col = c["hire_date"]
    trans_col = c["transdate"]

    # Batch-load residents
    resident_lookup: dict[int, Resident] = {
        r.id: r
        for r in db.session.query(Resident).filter(
            Resident.id.in_(resident_data.keys())
        )
    }

    for resident_id, data in sorted(
        resident_data.items(), key=lambda item: item[1]["name"]
    ):
        resident = resident_lookup.get(resident_id)
        if resident is None:
            continue

        row: list[object | None] = base_row.copy()
        row[hire_col] = resident.hire_date
        row[c["employee"]] = data["name"]
        row[c["lawson_id"]] = (
            str(resident.lawson_id) if resident.lawson_id is not None else None
        )
        row[c["hours"]] = round(data["total_overtime"], 2)

        if row[c["hours"]] == 0:
            continue  # Skip zero-hour rows
        ws.append(row)

        # Format the appended row
        r = ws.max_row
        for col0 in text_cols:
            ws.cell(row=r, column=col0 + 1).number_format = settings.text_format
        ws.cell(row=r, column=hire_col + 1).number_format = settings.date_format
        ws.cell(row=r, column=trans_col + 1).number_format = settings.date_format

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
