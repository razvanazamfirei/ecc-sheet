"""Tests for report utilities."""

import io
from datetime import date, time

import openpyxl

from backend.models import PayrollSettings, Resident, TimeEntry, db
from backend.report_utils import (
    aggregate_entries_by_resident,
    build_entries_query,
    generate_csv_content,
    generate_payroll_xlsx,
    get_resident_name,
)


class TestBuildEntriesQuery:
    """Tests for build_entries_query function."""

    def test_query_with_date_range(self, app):
        """Test building query with date range."""
        with app.app_context():
            start = date(2024, 1, 1)
            end = date(2024, 1, 31)
            query = build_entries_query(start, end, None)
            # Query should be built without error
            assert query is not None

    def test_query_with_resident_filter(self, app, sample_resident):
        """Test building query with resident filter."""
        with app.app_context():
            start = date(2024, 1, 1)
            end = date(2024, 1, 31)
            query = build_entries_query(start, end, sample_resident.id)
            assert query is not None


class TestGetResidentName:
    """Tests for get_resident_name function."""

    def test_returns_none_for_none_id(self, app):
        """Test that None is returned for None resident_id."""
        with app.app_context():
            assert get_resident_name(None) is None

    def test_returns_none_for_empty_string(self, app):
        """Test that None is returned for empty string resident_id."""
        with app.app_context():
            assert get_resident_name("") is None

    def test_returns_name_for_valid_id(self, app, sample_resident):
        """Test that name is returned for valid resident_id."""
        with app.app_context():
            name = get_resident_name(sample_resident.id)
            assert name == sample_resident.name

    def test_returns_none_for_invalid_id(self, app):
        """Test that None is returned for non-existent resident_id."""
        with app.app_context():
            assert get_resident_name(99999) is None


class TestAggregateEntriesByResident:
    """Tests for aggregate_entries_by_resident function."""

    def test_empty_entries(self, app):
        """Test aggregation with empty entries list."""
        with app.app_context():
            result = aggregate_entries_by_resident([])
            assert result == {}

    def test_single_entry(self, app, sample_time_entry):
        """Test aggregation with single entry."""
        with app.app_context():
            # Reload entry in current session
            entry = db.session.get(TimeEntry, sample_time_entry.id)
            assert entry is not None
            result = aggregate_entries_by_resident([entry])

            assert len(result) == 1
            assert entry.resident_id in result
            data = result[entry.resident_id]
            assert data["name"] == entry.resident.name
            assert len(data["entries"]) == 1
            assert data["total_overtime"] >= 0

    def test_multiple_entries_same_resident(self, app, sample_resident, sample_role):
        """Test aggregation with multiple entries for same resident."""
        with app.app_context():
            # Create multiple entries
            entries = []
            for i in range(3):
                entry = TimeEntry(
                    date=date(2024, 1, i + 1),
                    resident_id=sample_resident.id,
                    role_id=sample_role.id,
                    exit_time=time(19, 0),
                )
                db.session.add(entry)
                entries.append(entry)
            db.session.commit()

            # Reload entries
            entries = TimeEntry.query.filter(
                TimeEntry.resident_id == sample_resident.id
            ).all()

            result = aggregate_entries_by_resident(entries)

            assert len(result) == 1
            assert sample_resident.id in result
            assert len(result[sample_resident.id]["entries"]) == len(entries)

            # Cleanup
            for entry in entries:
                db.session.delete(entry)
            db.session.commit()

    def test_entry_without_exit_time(self, app, sample_resident, sample_role):
        """Test aggregation with entry missing exit time."""
        with app.app_context():
            entry = TimeEntry(
                date=date(2024, 1, 1),
                resident_id=sample_resident.id,
                role_id=sample_role.id,
                exit_time=None,
            )
            db.session.add(entry)
            db.session.commit()

            imported_entry = db.session.get(TimeEntry, entry.id)
            assert imported_entry is not None
            entries = [imported_entry]
            result = aggregate_entries_by_resident(entries)

            assert len(result) == 1
            # Exit time should be empty string
            assert not result[sample_resident.id]["entries"][0]["exit_time"]

            db.session.delete(entry)
            db.session.commit()


class TestGenerateCsvContent:
    """Tests for generate_csv_content function."""

    def test_empty_entries(self, app):
        """Test CSV generation with empty entries."""
        with app.app_context():
            csv_content = generate_csv_content([])
            lines = csv_content.strip().split("\n")
            assert len(lines) == 1  # Just header
            assert "Date" in lines[0]
            assert "Resident" in lines[0]

    def test_single_entry(self, app, sample_time_entry):
        """Test CSV generation with single entry."""
        with app.app_context():
            entry = db.session.get(TimeEntry, sample_time_entry.id)
            assert entry is not None
            csv_content = generate_csv_content([entry])

            lines = csv_content.strip().split("\n")
            assert len(lines) == 2  # Header + 1 entry
            assert entry.resident.name in lines[1]
            assert entry.role.name in lines[1]

    def test_entry_without_exit_time(self, app, sample_resident, sample_role):
        """Test CSV generation with entry missing exit time."""
        with app.app_context():
            entry = TimeEntry(
                date=date(2024, 1, 1),
                resident_id=sample_resident.id,
                role_id=sample_role.id,
                exit_time=None,
            )
            db.session.add(entry)
            db.session.commit()

            imported_entry = db.session.get(TimeEntry, entry.id)
            assert imported_entry is not None
            entries = [imported_entry]
            csv_content = generate_csv_content(entries)

            # Should not raise error
            assert csv_content is not None
            lines = csv_content.strip().split("\n")
            assert len(lines) == 2

            db.session.delete(entry)
            db.session.commit()


class TestGeneratePayrollXlsx:
    """Tests for generate_payroll_xlsx function."""

    def test_returns_bytes(self, app, sample_resident):
        """Test that the function returns bytes."""
        with app.app_context():
            settings = PayrollSettings.get_or_create()
            orig_program = settings.program
            orig_company = settings.company
            orig_label_suffix = settings.label_suffix
            try:
                settings.program = "M1300"
                settings.company = "UPHS"
                settings.label_suffix = "ECA"
                db.session.commit()

                resident_data = {
                    sample_resident.id: {
                        "name": sample_resident.name,
                        "entries": [],
                        "total_overtime": 2.5,
                    }
                }
                result = generate_payroll_xlsx(
                    resident_data,
                    date(2026, 1, 1),
                    date(2026, 1, 31),
                    settings,
                )
                assert isinstance(result, bytes)
                assert len(result) > 0
            finally:
                settings.program = orig_program
                settings.company = orig_company
                settings.label_suffix = orig_label_suffix
                db.session.commit()

    def test_xlsx_valid_workbook(self, app, sample_resident):
        """Test that returned bytes form a valid xlsx workbook."""
        with app.app_context():
            settings = PayrollSettings.get_or_create()
            settings.label_suffix = "ECA"
            db.session.commit()

            resident_data = {
                sample_resident.id: {
                    "name": sample_resident.name,
                    "entries": [],
                    "total_overtime": 1.0,
                }
            }
            result = generate_payroll_xlsx(
                resident_data,
                date(2026, 1, 1),
                date(2026, 1, 31),
                settings,
            )
            wb = openpyxl.load_workbook(io.BytesIO(result))
            ws = wb.active
            assert ws.cell(row=1, column=1).value == "Program"
            assert ws.cell(row=1, column=9).value == "Hours"

    def test_excludes_residents_without_lawson_id(self, app, sample_resident):
        """Test that residents without lawson_id are excluded."""
        with app.app_context():
            resident = db.session.get(Resident, sample_resident.id)
            resident.lawson_id = None
            db.session.commit()

            settings = PayrollSettings.get_or_create()
            settings.label_suffix = "ECA"
            db.session.commit()

            resident_data = {
                resident.id: {
                    "name": resident.name,
                    "entries": [],
                    "total_overtime": 3.0,
                }
            }
            result = generate_payroll_xlsx(
                resident_data,
                date(2026, 1, 1),
                date(2026, 1, 31),
                settings,
            )
            wb = openpyxl.load_workbook(io.BytesIO(result))
            ws = wb.active
            assert ws.max_row == 1  # Only header

    def test_includes_residents_with_lawson_id(self, app, sample_resident):
        """Test that residents with lawson_id appear as data rows."""
        with app.app_context():
            resident = db.session.get(Resident, sample_resident.id)
            resident.lawson_id = 12345
            db.session.commit()

            settings = PayrollSettings.get_or_create()
            settings.label_suffix = "ECA"
            db.session.commit()

            resident_data = {
                resident.id: {
                    "name": resident.name,
                    "entries": [],
                    "total_overtime": 2.0,
                }
            }
            result = generate_payroll_xlsx(
                resident_data,
                date(2026, 1, 1),
                date(2026, 1, 31),
                settings,
            )
            wb = openpyxl.load_workbook(io.BytesIO(result))
            ws = wb.active
            assert ws.max_row == 2  # Header + 1 data row
            assert ws.cell(row=2, column=9).value == 2  # Hours

    def test_col_ab_note_format(self, app, sample_resident):
        """Test that col AB contains '{MON} {label_suffix}'."""
        with app.app_context():
            resident = db.session.get(Resident, sample_resident.id)
            resident.lawson_id = 99999
            db.session.commit()

            settings = PayrollSettings.get_or_create()
            settings.label_suffix = "ECA"
            db.session.commit()

            resident_data = {
                resident.id: {
                    "name": resident.name,
                    "entries": [],
                    "total_overtime": 1.5,
                }
            }
            result = generate_payroll_xlsx(
                resident_data,
                date(2026, 3, 1),
                date(2026, 3, 31),
                settings,
            )
            wb = openpyxl.load_workbook(io.BytesIO(result))
            ws = wb.active
            assert ws.cell(row=2, column=28).value == "MAR ECA"

    def test_transdate_is_end_date(self, app, sample_resident):
        """Test that Transdate column (N = col 14) contains end_date as MM/DD/YYYY."""
        with app.app_context():
            resident = db.session.get(Resident, sample_resident.id)
            resident.lawson_id = 11111
            db.session.commit()

            settings = PayrollSettings.get_or_create()
            settings.label_suffix = "ECA"
            db.session.commit()

            end = date(2026, 2, 28)
            resident_data = {
                resident.id: {
                    "name": resident.name,
                    "entries": [],
                    "total_overtime": 0.5,
                }
            }
            result = generate_payroll_xlsx(
                resident_data,
                date(2026, 2, 1),
                end,
                settings,
            )
            wb = openpyxl.load_workbook(io.BytesIO(result))
            ws = wb.active
            cell_value = ws.cell(row=2, column=14).value
            assert cell_value == end.strftime("%m/%d/%Y")


class TestPayrollSettingsModel:
    """Tests for PayrollSettings model."""

    def test_get_or_create_returns_instance(self, app):
        """Test that get_or_create always returns a settings object."""
        with app.app_context():
            settings = PayrollSettings.get_or_create()
            assert settings is not None
            assert settings.id is not None

    def test_get_or_create_idempotent(self, app):
        """Test that calling get_or_create twice returns the same row."""
        with app.app_context():
            s1 = PayrollSettings.get_or_create()
            s2 = PayrollSettings.get_or_create()
            assert s1.id == s2.id

    def test_fields_persist(self, app):
        """Test that settings fields can be saved and retrieved."""
        with app.app_context():
            settings = PayrollSettings.get_or_create()
            settings.program = "M9999"
            settings.company = "TEST"
            settings.batch = 42
            settings.label_suffix = "XYZ"
            db.session.commit()

            reloaded = PayrollSettings.query.first()
            assert reloaded.program == "M9999"
            assert reloaded.company == "TEST"
            assert reloaded.batch == 42
            assert reloaded.label_suffix == "XYZ"

            # Reset to avoid affecting other tests
            settings.program = None
            settings.company = None
            settings.batch = None
            settings.label_suffix = None
            db.session.commit()
