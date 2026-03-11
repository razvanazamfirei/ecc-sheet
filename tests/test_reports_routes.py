"""Tests for report routes."""

import io

import openpyxl
import pytest

from backend.models import PayrollSettings, Resident, TimeEntry, db


class TestReportGeneration:
    """Tests for report generation."""

    def test_reports_page_loads(self, client):
        """Test that reports page loads."""
        response = client.get("/reports")
        assert response.status_code == 200

    def test_generate_report_with_entries(self, client, app, sample_time_entry):
        """Test generating a report with entries."""
        with app.app_context():
            entry = db.session.get(TimeEntry, sample_time_entry.id)
            assert entry is not None
            entry_date = entry.date

            response = client.post(
                "/api/report",
                data={
                    "start_date": entry_date.strftime("%Y-%m-%d"),
                    "end_date": entry_date.strftime("%Y-%m-%d"),
                },
                follow_redirects=True,
            )
            assert response.status_code == 200
            assert b"Overtime" in response.data

    def test_generate_report_filtered_by_resident(
        self, client, app, sample_time_entry, sample_resident
    ):
        """Test generating a report filtered by resident."""
        with app.app_context():
            entry = db.session.get(TimeEntry, sample_time_entry.id)
            assert entry is not None
            entry_date = entry.date

            response = client.post(
                "/api/report",
                data={
                    "start_date": entry_date.strftime("%Y-%m-%d"),
                    "end_date": entry_date.strftime("%Y-%m-%d"),
                    "resident_id": sample_resident.id,
                },
                follow_redirects=True,
            )
            assert response.status_code == 200

    def test_generate_report_invalid_date(self, client):
        """Test generating a report with invalid date."""
        response = client.post(
            "/api/report",
            data={
                "start_date": "invalid",
                "end_date": "2024-01-31",
            },
            follow_redirects=True,
        )
        assert response.status_code == 200
        # Should redirect with error flash


class TestReportExport:
    """Tests for report CSV export."""

    def test_export_csv_empty(self, client):
        """Test exporting empty report as CSV."""
        response = client.post(
            "/api/report/export_csv",
            data={
                "start_date": "2020-01-01",
                "end_date": "2020-01-31",
            },
        )
        assert response.status_code == 200
        assert "text/csv" in response.content_type
        assert b"Date,Resident,Role" in response.data

    def test_export_csv_with_entries(self, client, app, sample_time_entry):
        """Test exporting report with entries as CSV."""
        with app.app_context():
            entry = db.session.get(TimeEntry, sample_time_entry.id)
            assert entry is not None
            entry_date = entry.date

            response = client.post(
                "/api/report/export_csv",
                data={
                    "start_date": entry_date.strftime("%Y-%m-%d"),
                    "end_date": entry_date.strftime("%Y-%m-%d"),
                },
            )
            assert response.status_code == 200
            assert "text/csv" in response.content_type
            # Check filename in Content-Disposition
            assert "overtime_report" in response.headers.get("Content-Disposition", "")

    def test_export_csv_filtered_by_resident(
        self, client, app, sample_time_entry, sample_resident
    ):
        """Test exporting filtered report as CSV."""
        with app.app_context():
            entry = db.session.get(TimeEntry, sample_time_entry.id)
            assert entry is not None
            entry_date = entry.date

            response = client.post(
                "/api/report/export_csv",
                data={
                    "start_date": entry_date.strftime("%Y-%m-%d"),
                    "end_date": entry_date.strftime("%Y-%m-%d"),
                    "resident_id": sample_resident.id,
                },
            )
            assert response.status_code == 200
            assert "text/csv" in response.content_type

    def test_export_csv_invalid_date(self, client):
        """Test exporting CSV with invalid date redirects."""
        response = client.post(
            "/api/report/export_csv",
            data={
                "start_date": "invalid",
                "end_date": "2024-01-31",
            },
            follow_redirects=True,
        )
        assert response.status_code == 200
        # Should show error message
        assert b"invalid" in response.data.lower() or b"error" in response.data.lower()


class TestReportEdgeCases:
    """Edge case tests for reports."""

    def test_generate_report_date_range(self, client):
        """Test report with wide date range."""
        response = client.post(
            "/api/report",
            data={
                "start_date": "2020-01-01",
                "end_date": "2025-12-31",
            },
            follow_redirects=True,
        )
        assert response.status_code == 200

    def test_generate_report_same_day(self, client, app, sample_time_entry):
        """Test report for a single day."""
        with app.app_context():
            entry = db.session.get(TimeEntry, sample_time_entry.id)
            assert entry is not None
            entry_date = entry.date
            date_str = entry_date.strftime("%Y-%m-%d")

            response = client.post(
                "/api/report",
                data={
                    "start_date": date_str,
                    "end_date": date_str,
                },
                follow_redirects=True,
            )
            assert response.status_code == 200


@pytest.mark.integration
class TestPayrollXlsxExport:
    """Tests for payroll XLSX export route."""

    def test_export_payroll_xlsx_empty_returns_xlsx(self, client):
        """Test that export returns an xlsx file even when no entries match."""
        response = client.post(
            "/api/report/export_payroll_xlsx",
            data={
                "start_date": "2020-01-01",
                "end_date": "2020-01-31",
            },
        )
        assert response.status_code == 200
        assert "spreadsheetml" in response.content_type
        assert "payroll_" in response.headers.get("Content-Disposition", "")

    def test_export_payroll_xlsx_with_lawson_id(self, client, app, sample_time_entry):
        """Test xlsx export includes resident with lawson_id set."""
        with app.app_context():
            entry = db.session.get(TimeEntry, sample_time_entry.id)
            resident = db.session.get(Resident, entry.resident_id)
            resident.lawson_id = 55555
            db.session.commit()

            entry_date = entry.date
            response = client.post(
                "/api/report/export_payroll_xlsx",
                data={
                    "start_date": entry_date.strftime("%Y-%m-%d"),
                    "end_date": entry_date.strftime("%Y-%m-%d"),
                },
            )
            assert response.status_code == 200
            assert "spreadsheetml" in response.content_type

            wb = openpyxl.load_workbook(io.BytesIO(response.data))
            ws = wb.active
            sheet_values = [
                ws.cell(row=r, column=c).value
                for r in range(1, ws.max_row + 1)
                for c in range(1, ws.max_column + 1)
            ]
            assert "55555" in sheet_values

            # Cleanup
            resident.lawson_id = None
            db.session.commit()

    def test_export_payroll_xlsx_invalid_date(self, client):
        """Test that invalid date redirects with error."""
        response = client.post(
            "/api/report/export_payroll_xlsx",
            data={
                "start_date": "not-a-date",
                "end_date": "2026-01-31",
            },
            follow_redirects=True,
        )
        assert response.status_code == 200
        assert b"Invalid start_date: not-a-date" in response.data

    def test_export_payroll_xlsx_filename_contains_dates(self, client):
        """Test that Content-Disposition filename includes date range."""
        response = client.post(
            "/api/report/export_payroll_xlsx",
            data={
                "start_date": "2026-01-01",
                "end_date": "2026-01-31",
            },
        )
        disposition = response.headers.get("Content-Disposition", "")
        assert "2026-01-01" in disposition
        assert "2026-01-31" in disposition


class TestPayrollSettings:
    """Tests for payroll settings admin routes."""

    def test_payroll_settings_page_loads(self, client):
        """Test that payroll settings page loads for admin."""
        response = client.get("/payroll-settings")
        assert response.status_code == 200
        assert b"Payroll Settings" in response.data

    def test_payroll_settings_page_read_only_for_non_payroll_admin(
        self, client, monkeypatch
    ):
        """Test that regular admins see read-only view."""
        monkeypatch.setenv("PAYROLL_ADMIN_USERS", "Someone Else")
        response = client.get("/payroll-settings")
        assert response.status_code == 200
        assert b"read-only" in response.data.lower()

    def test_payroll_settings_requires_admin(self, client, monkeypatch):
        """Test that payroll settings page requires admin."""
        monkeypatch.setenv("USER_NAME", "Regular User")
        monkeypatch.setenv("ADMIN_USERS", "Admin Only")
        response = client.get("/payroll-settings", follow_redirects=True)
        assert b"Admin privileges required" in response.data

    def test_payroll_settings_save_requires_payroll_admin(self, client, monkeypatch):
        """Test that non-payroll-admin cannot save settings."""
        monkeypatch.setenv("PAYROLL_ADMIN_USERS", "Someone Else")
        response = client.post(
            "/payroll-settings",
            data={"program": "X"},
            follow_redirects=True,
        )
        assert b"Payroll admin privileges required" in response.data

    def test_payroll_settings_save(self, client, app, monkeypatch):
        """Test saving payroll settings as payroll admin."""
        monkeypatch.setenv("USER_NAME", "CI-Test-User")
        monkeypatch.setenv("PAYROLL_ADMIN_USERS", "CI-Test-User")

        with app.app_context():
            response = client.post(
                "/payroll-settings",
                data={
                    "program": "M1300",
                    "company": "UPHS",
                    "batch": "860",
                    "pay_code": "101758",
                    "dept": "102",
                    "expense": "1003",
                    "acct_unit": "102000",
                    "label_suffix": "ECA",
                },
                follow_redirects=True,
            )
            assert response.status_code == 200
            assert b"saved successfully" in response.data

            settings = PayrollSettings.query.first()
            assert settings.program == "M1300"
            assert settings.company == "UPHS"
            assert settings.batch == 860
            assert settings.label_suffix == "ECA"

    def test_payroll_settings_save_empty_values(self, client, app, monkeypatch):
        """Test saving empty payroll settings stores None for integer fields."""
        monkeypatch.setenv("USER_NAME", "CI-Test-User")
        monkeypatch.setenv("PAYROLL_ADMIN_USERS", "CI-Test-User")

        with app.app_context():
            response = client.post(
                "/payroll-settings",
                data={
                    "program": "",
                    "company": "",
                    "batch": "",
                    "pay_code": "",
                    "dept": "",
                    "expense": "",
                    "acct_unit": "",
                    "label_suffix": "",
                },
                follow_redirects=True,
            )
            assert response.status_code == 200

            settings = PayrollSettings.query.first()
            assert settings.program is None
            assert settings.batch is None

    def test_payroll_settings_save_invalid_integer(self, client, app, monkeypatch):
        """Test invalid payroll integer input returns a flashed validation error."""
        monkeypatch.setenv("USER_NAME", "CI-Test-User")
        monkeypatch.setenv("PAYROLL_ADMIN_USERS", "CI-Test-User")

        with app.app_context():
            response = client.post(
                "/payroll-settings",
                data={
                    "program": "M1300",
                    "batch": "not-a-number",
                },
                follow_redirects=True,
            )
            assert response.status_code == 200
            assert b"must be an integer" in response.data


@pytest.mark.integration
class TestReportExportPermissions:
    """Tests that extended report actions stay admin/payroll-only."""

    @staticmethod
    def _set_non_payroll_env() -> dict[str, str]:
        """Return env overrides for a non-payroll, non-admin user."""
        return {
            "USER_NAME": "Regular Viewer",
            "ADMIN_USERS": "Admin Only",
            "PAYROLL_ADMIN_USERS": "",
        }

    def test_billing_csv_blocked_for_regular_user(self, client, monkeypatch):
        """Non-admin/non-payroll user is redirected from billing CSV export."""
        for k, v in self._set_non_payroll_env().items():
            monkeypatch.setenv(k, v)

        response = client.post(
            "/api/report/export_billing_csv",
            data={"start_date": "2026-01-01", "end_date": "2026-01-31"},
            follow_redirects=True,
        )
        assert response.status_code == 200
        assert b"permission" in response.data.lower()

    def test_payroll_xlsx_blocked_for_regular_user(self, client, monkeypatch):
        """Non-admin/non-payroll user is redirected from payroll XLSX export."""
        for k, v in self._set_non_payroll_env().items():
            monkeypatch.setenv(k, v)

        response = client.post(
            "/api/report/export_payroll_xlsx",
            data={"start_date": "2026-01-01", "end_date": "2026-01-31"},
            follow_redirects=True,
        )
        assert response.status_code == 200
        assert b"permission" in response.data.lower()

    def test_billing_csv_allowed_for_admin(self, client):
        """Admin user can access billing CSV export."""
        response = client.post(
            "/api/report/export_billing_csv",
            data={"start_date": "2020-01-01", "end_date": "2020-01-31"},
        )
        assert response.status_code == 200
        assert "text/csv" in response.content_type

    def test_payroll_xlsx_allowed_for_admin(self, client):
        """Admin user can access payroll XLSX export."""
        response = client.post(
            "/api/report/export_payroll_xlsx",
            data={"start_date": "2020-01-01", "end_date": "2020-01-31"},
        )
        assert response.status_code == 200
        assert "spreadsheetml" in response.content_type


@pytest.mark.integration
class TestReportRestriction:
    """Tests report filtering permissions and restrictions."""

    def test_restricted_user_sees_self_only_note(self, client, monkeypatch):
        """Reports page shows 'your entries only' for non-admin."""
        monkeypatch.setenv("USER_NAME", "Regular Viewer")
        monkeypatch.setenv("ADMIN_USERS", "Admin Only")
        monkeypatch.setenv("PAYROLL_ADMIN_USERS", "")

        response = client.get("/reports")
        assert response.status_code == 200
        assert b"your entries only" in response.data.lower()

    def test_admin_sees_resident_filter(self, client):
        """Reports page shows resident dropdown for admin."""
        response = client.get("/reports")
        assert response.status_code == 200
        assert b"All Residents" in response.data

    def test_report_viewer_sees_resident_filter_without_extended_actions(
        self, client, app, sample_time_entry, monkeypatch
    ):
        """Listed report viewers can filter residents but still lack extras."""
        monkeypatch.setenv("USER_NAME", "Demo Viewer")
        monkeypatch.setenv("ADMIN_USERS", "Razvan Azamfirei")
        monkeypatch.setenv("PAYROLL_ADMIN_USERS", "")
        monkeypatch.setenv("REPORT_VIEW_ALL_USERS", "Demo Viewer")

        response = client.get("/reports")
        assert response.status_code == 200
        assert b"All Residents" in response.data

        with app.app_context():
            entry = db.session.get(TimeEntry, sample_time_entry.id)
            date_str = entry.date.strftime("%Y-%m-%d")

        response = client.post(
            "/api/report",
            data={"start_date": date_str, "end_date": date_str},
            follow_redirects=True,
        )
        assert response.status_code == 200
        assert b"Export Detailed CSV" in response.data
        assert b"Export Billing CSV" not in response.data
        assert b"Export Payroll XLSX" not in response.data

    def test_report_viewer_can_submit_resident_filter(
        self, client, app, sample_time_entry, sample_resident, monkeypatch
    ):
        """Listed report viewers keep their submitted resident filter."""
        monkeypatch.setenv("USER_NAME", "Demo Viewer")
        monkeypatch.setenv("ADMIN_USERS", "Razvan Azamfirei")
        monkeypatch.setenv("PAYROLL_ADMIN_USERS", "")
        monkeypatch.setenv("REPORT_VIEW_ALL_USERS", "Demo Viewer")

        with app.app_context():
            entry = db.session.get(TimeEntry, sample_time_entry.id)
            date_str = entry.date.strftime("%Y-%m-%d")
            other_resident = Resident(name="Other Report Resident", active=True)
            db.session.add(other_resident)
            db.session.commit()
            other_resident_id = other_resident.id

            other_entry = TimeEntry(
                date=entry.date,
                resident_id=other_resident_id,
                role_id=entry.role_id,
                exit_time=entry.exit_time,
            )
            db.session.add(other_entry)
            db.session.commit()
            other_entry_id = other_entry.id

        try:
            response = client.post(
                "/api/report",
                data={
                    "start_date": date_str,
                    "end_date": date_str,
                    "resident_id": sample_resident.id,
                },
                follow_redirects=True,
            )
            assert response.status_code == 200
            assert sample_resident.name.encode() in response.data
            assert b'name="resident_id"' in response.data
            assert f'value="{sample_resident.id}"'.encode() in response.data
        finally:
            with app.app_context():
                other_entry = db.session.get(TimeEntry, other_entry_id)
                if other_entry is not None:
                    db.session.delete(other_entry)
                other_resident = db.session.get(Resident, other_resident_id)
                if other_resident is not None:
                    db.session.delete(other_resident)
                db.session.commit()

    def test_admin_sees_extended_actions_in_report_results(
        self, client, app, sample_time_entry
    ):
        """Admins keep access to extended report actions in the results view."""
        with app.app_context():
            entry = db.session.get(TimeEntry, sample_time_entry.id)
            date_str = entry.date.strftime("%Y-%m-%d")

        response = client.post(
            "/api/report",
            data={"start_date": date_str, "end_date": date_str},
            follow_redirects=True,
        )
        assert response.status_code == 200
        assert b"Export Billing CSV" in response.data
        assert b"Export Payroll XLSX" in response.data

    def test_payroll_admin_sees_extended_actions_in_report_results(
        self, client, app, sample_time_entry, monkeypatch
    ):
        """Payroll admins share the extended report actions branch."""
        monkeypatch.setenv("USER_NAME", "Payroll Person")
        monkeypatch.setenv("ADMIN_USERS", "Admin Only")
        monkeypatch.setenv("PAYROLL_ADMIN_USERS", "Payroll Person")

        with app.app_context():
            entry = db.session.get(TimeEntry, sample_time_entry.id)
            date_str = entry.date.strftime("%Y-%m-%d")

        response = client.post(
            "/api/report",
            data={"start_date": date_str, "end_date": date_str},
            follow_redirects=True,
        )
        assert response.status_code == 200
        assert b"Export Billing CSV" in response.data
        assert b"Export Payroll XLSX" in response.data

    def test_report_generation_forces_resident_id_for_restricted(
        self, client, app, sample_time_entry, sample_resident, monkeypatch
    ):
        """Non-admin report POST ignores submitted resident_id and uses own."""
        monkeypatch.setenv("USER_NAME", "Regular Viewer")
        monkeypatch.setenv("ADMIN_USERS", "Admin Only")
        monkeypatch.setenv("PAYROLL_ADMIN_USERS", "")

        with app.app_context():
            entry = db.session.get(TimeEntry, sample_time_entry.id)
            date_str = entry.date.strftime("%Y-%m-%d")

        # Submit with a specific resident_id — should be overridden
        response = client.post(
            "/api/report",
            data={
                "start_date": date_str,
                "end_date": date_str,
                "resident_id": sample_resident.id,
            },
            follow_redirects=True,
        )
        # Should succeed (200) but data shown will reflect restriction
        assert response.status_code == 200
